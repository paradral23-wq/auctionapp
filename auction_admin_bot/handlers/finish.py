"""
handlers/finish.py
Отчёт по завершённому лоту + экспорт в Excel / CSV.
"""
import csv
import io
import logging
from datetime import timezone, timedelta

from aiogram import F, Router
from aiogram.types import BufferedInputFile, CallbackQuery

from db.queries import get_bid_count, get_lot, get_lot_bids, get_unique_bidder_count
from keyboards.inline import kb_back_to_main, kb_report_actions, kb_winner
from utils.formatting import fmt_aed, report_text
from utils.guards import admin_only_callback

logger = logging.getLogger(__name__)
router = Router()

MSK = timezone(timedelta(hours=3))


def _winner_str(lot) -> str:
    if lot.winner_username:
        return f"@{lot.winner_username}"
    if lot.winner_user_id and lot.winner_user_id != 0:
        return f"id{lot.winner_user_id}"
    return "—"


def _sold(lot) -> bool:
    return bool(lot.winner_user_id and lot.winner_user_id != 0)


def _final_discount(lot) -> str:
    if lot.market_price and lot.final_price:
        pct = round((1 - lot.final_price / lot.market_price) * 100)
        return f"{pct}%"
    return "—"


def _drop_from_start(lot) -> str:
    if lot.final_price and lot.start_price and lot.final_price < lot.start_price:
        return fmt_aed(lot.start_price - lot.final_price)
    return "—"


# ── Отчёт (текстовый) ─────────────────────────────────────────

@router.callback_query(F.data.startswith("win:report:"))
async def cb_report(callback: CallbackQuery):
    if not await admin_only_callback(callback):
        return
    lot_id = int(callback.data.split(":")[2])
    lot = await get_lot(lot_id)
    if not lot:
        await callback.answer("Лот не найден.", show_alert=True)
        return
    bid_count  = await get_bid_count(lot_id)
    user_count = await get_unique_bidder_count(lot_id)
    await callback.message.edit_text(
        report_text(lot, bid_count, user_count),
        reply_markup=kb_report_actions(lot_id, back_cb="lots:finished"),
        parse_mode="HTML",
    )
    await callback.answer()


# ── Экспорт Excel ─────────────────────────────────────────────

@router.callback_query(F.data.startswith("export:xlsx:"))
async def cb_export_xlsx(callback: CallbackQuery):
    if not await admin_only_callback(callback):
        return

    lot_id = int(callback.data.split(":")[2])
    lot = await get_lot(lot_id)
    if not lot:
        await callback.answer("Лот не найден.", show_alert=True)
        return

    await callback.answer("Формирую Excel...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        await callback.message.answer(
            "❌ openpyxl не установлен.\n<code>pip install openpyxl</code>",
            parse_mode="HTML",
        )
        return

    bids       = await get_lot_bids(lot_id, limit=500)
    bid_count  = await get_bid_count(lot_id)
    user_count = await get_unique_bidder_count(lot_id)
    sold       = _sold(lot)

    wb = openpyxl.Workbook()

    # ── Стили ─────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1a3a5c")
    section_fill = PatternFill("solid", fgColor="2d4a6e")
    label_font   = Font(bold=True, size=10)
    value_font   = Font(size=10)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left", vertical="center")
    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, text):
        cell.value     = text
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = section_fill
        cell.alignment = center
        cell.border    = border

    def row_kv(ws, label, value, r):
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.font = label_font; lc.border = border; lc.alignment = left
        vc.font = value_font; vc.border = border; vc.alignment = left

    # ── Лист 1: Сводка ────────────────────────────────────────
    ws = wb.active
    ws.title = "Сводка"

    # Заголовок
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value     = f"{lot.emoji} {lot.title}  |  {lot.lot_code}"
    t.font      = Font(bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor="0d2137")
    t.alignment = center
    ws.row_dimensions[1].height = 28

    r = 2
    # Объект
    ws.merge_cells(f"A{r}:B{r}"); hdr(ws[f"A{r}"], "ОБЪЕКТ"); r += 1
    row_kv(ws, "Тип",          lot.property_type or "—",  r); r += 1
    row_kv(ws, "Площадь",      f"{lot.area_sqft:,} sqft" if lot.area_sqft else "—", r); r += 1
    row_kv(ws, "Этаж",         lot.floor_level or "—",    r); r += 1
    row_kv(ws, "Вид",          lot.view_type or "—",      r); r += 1
    row_kv(ws, "Парковка",     str(lot.parking_spots) if lot.parking_spots is not None else "—", r); r += 1
    row_kv(ws, "Статус",       lot.property_status or "—", r); r += 1
    row_kv(ws, "Описание",     lot.description or "—",    r); r += 1

    # Параметры аукциона
    ws.merge_cells(f"A{r}:B{r}"); hdr(ws[f"A{r}"], "ПАРАМЕТРЫ АУКЦИОНА (AED)"); r += 1
    if lot.purchase_price:
        row_kv(ws, "Цена покупки",  lot.purchase_price, r); r += 1
    if lot.market_price:
        row_kv(ws, "Рыночная цена", lot.market_price,   r); r += 1
    if lot.discount_pct:
        row_kv(ws, "Скидка к рынку", f"{lot.discount_pct}%", r); r += 1
    row_kv(ws, "Старт аукциона", lot.start_price,        r); r += 1
    row_kv(ws, "Шаг снижения",   lot.bid_step,           r); r += 1
    row_kv(ws, "Интервал",       f"{lot.price_drop_interval_minutes} мин" if lot.price_drop_interval_minutes else "—", r); r += 1

    # Итог
    ws.merge_cells(f"A{r}:B{r}"); hdr(ws[f"A{r}"], "ИТОГ"); r += 1
    if sold:
        row_kv(ws, "Статус",             "Продан",          r); r += 1
        row_kv(ws, "Финальная цена",     lot.final_price,   r); r += 1
        row_kv(ws, "Победитель",         _winner_str(lot),  r); r += 1
        row_kv(ws, "Снижение от старта", _drop_from_start(lot), r); r += 1
        row_kv(ws, "Скидка от рынка",    _final_discount(lot),  r); r += 1
    else:
        row_kv(ws, "Статус",       "Не состоялся",           r); r += 1
        row_kv(ws, "Цена дошла до", fmt_aed(lot.current_price), r); r += 1
        row_kv(ws, "Победитель",   "—",                       r); r += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 32

    # ── Лист 2: История снижений (ставки = снижения цены) ─────
    ws2 = wb.create_sheet("История снижений")
    headers = ["#", "Дата/Время (МСК)", "Цена (AED)", "Покупатель", "Telegram ID"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = border

    for i, bid in enumerate(bids, 1):
        dt = bid.created_at.astimezone(MSK).strftime("%d.%m.%Y %H:%M:%S")
        buyer = f"@{bid.username}" if bid.username else (f"id{bid.user_id}" if bid.user_id else "—")
        is_winner = sold and bid.user_id == lot.winner_user_id
        vals = [i, dt, bid.amount, buyer if is_winner else "—", bid.user_id if is_winner else "—"]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=i + 1, column=col, value=val)
            c.border = border
            c.alignment = Alignment(vertical="center")
            if is_winner:
                c.font = Font(bold=True, color="1a5c1a", size=10)
            else:
                c.font = Font(size=10)

    for col, w in enumerate([5, 22, 18, 22, 16], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    status_str = f"Продан за {fmt_aed(lot.final_price)}" if sold else "Не состоялся"
    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename=f"report_{lot.lot_code}.xlsx"),
        caption=(
            f"📊 <b>{lot.lot_code}</b> — {lot.emoji} {lot.title}\n"
            f"{status_str}"
        ),
        parse_mode="HTML",
    )
    logger.info(f"Excel exported: lot {lot_id}")


# ── Экспорт CSV ───────────────────────────────────────────────

@router.callback_query(F.data.startswith("export:csv:"))
async def cb_export_csv(callback: CallbackQuery):
    if not await admin_only_callback(callback):
        return

    lot_id = int(callback.data.split(":")[2])
    lot = await get_lot(lot_id)
    if not lot:
        await callback.answer("Лот не найден.", show_alert=True)
        return

    await callback.answer("Формирую CSV...")

    bids       = await get_lot_bids(lot_id, limit=500)
    bid_count  = await get_bid_count(lot_id)
    user_count = await get_unique_bidder_count(lot_id)
    sold       = _sold(lot)

    buf = io.StringIO()
    w = csv.writer(buf)

    # Объект
    w.writerow(["=== ОБЪЕКТ ==="])
    w.writerow(["Код", lot.lot_code])
    w.writerow(["Название", lot.title])
    w.writerow(["Тип", lot.property_type or ""])
    w.writerow(["Площадь sqft", lot.area_sqft or ""])
    w.writerow(["Этаж", lot.floor_level or ""])
    w.writerow(["Вид", lot.view_type or ""])
    w.writerow(["Парковка", lot.parking_spots if lot.parking_spots is not None else ""])
    w.writerow(["Статус объекта", lot.property_status or ""])
    w.writerow(["Описание", lot.description or ""])
    w.writerow([])

    # Параметры аукциона
    w.writerow(["=== ПАРАМЕТРЫ АУКЦИОНА ==="])
    w.writerow(["Цена покупки AED", lot.purchase_price or ""])
    w.writerow(["Рыночная цена AED", lot.market_price or ""])
    w.writerow(["Скидка к рынку %", lot.discount_pct or ""])
    w.writerow(["Старт аукциона AED", lot.start_price])
    w.writerow(["Шаг снижения AED", lot.bid_step])
    w.writerow(["Интервал мин", lot.price_drop_interval_minutes or ""])
    w.writerow([])

    # Итог
    w.writerow(["=== ИТОГ ==="])
    if sold:
        w.writerow(["Статус", "Продан"])
        w.writerow(["Финальная цена AED", lot.final_price])
        w.writerow(["Победитель", _winner_str(lot)])
        w.writerow(["Снижение от старта AED", lot.start_price - (lot.final_price or 0) if lot.final_price else ""])
        w.writerow(["Скидка от рынка %", _final_discount(lot)])
    else:
        w.writerow(["Статус", "Не состоялся"])
        w.writerow(["Цена дошла до AED", lot.current_price])
        w.writerow(["Победитель", "—"])
    w.writerow([])

    # История снижений
    w.writerow(["=== ИСТОРИЯ СНИЖЕНИЙ ==="])
    w.writerow(["#", "Дата/Время МСК", "Цена AED", "Покупатель", "Telegram ID"])
    for i, bid in enumerate(bids, 1):
        dt = bid.created_at.astimezone(MSK).strftime("%d.%m.%Y %H:%M:%S")
        is_winner = sold and bid.user_id == lot.winner_user_id
        buyer = (_winner_str(lot) if is_winner else "—")
        uid   = (str(bid.user_id) if is_winner else "—")
        w.writerow([i, dt, bid.amount, buyer, uid])

    content = buf.getvalue().encode("utf-8-sig")

    status_str = f"Продан за {fmt_aed(lot.final_price)}" if sold else "Не состоялся"
    await callback.message.answer_document(
        BufferedInputFile(content, filename=f"report_{lot.lot_code}.csv"),
        caption=(
            f"📊 <b>{lot.lot_code}</b> — {lot.emoji} {lot.title}\n"
            f"{status_str}"
        ),
        parse_mode="HTML",
    )
    logger.info(f"CSV exported: lot {lot_id}")
